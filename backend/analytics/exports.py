
"""
Exports Excel et CSV des statistiques agrégées de Résumé+.

Seules des statistiques agrégées sont exportées (aucune donnée personnelle).
L'export respecte la période et les filtres sélectionnés (status, type,
service_id sont transmis aux agrégations).
"""
import csv
import io

from django.http import HttpResponse

from . import services

SECTION_LABELS = {
    'overview': 'Vue générale',
    'users': 'Utilisateurs',
    'summaries': 'Résumés',
    'qcm': 'QCM',
    'transactions': 'Transactions',
    'purchases': 'Achats de résumés',
    'subscriptions': 'Abonnements',
    'revenue': 'Revenus',
}

# Métriques top-level par section : (libellé, chemin 'clé1.clé2', clé d'arrondi)
_METRICS = {
    'overview': [
        ('Utilisateurs (total)', 'users.total'),
        ('Nouveaux utilisateurs (période)', 'users.new_in_period'),
        ('Résumés générés (total)', 'summaries.total'),
        ('Résumés générés (période)', 'summaries.in_period'),
        ('QCM générés (total)', 'qcm.total'),
        ('QCM générés (période)', 'qcm.in_period'),
        ('Transactions (période)', 'transactions.total'),
        ('  dont lancées', 'transactions.launched'),
        ('  dont réussies', 'transactions.succeeded'),
        ('  dont échouées', 'transactions.failed'),
        ('  dont annulées', 'transactions.cancelled'),
        ('Résumés achetés (total)', 'purchases.total'),
        ('Résumés achetés (période)', 'purchases.in_period'),
        ('Abonnements actifs', 'subscriptions.active'),
        ('Nouveaux abonnements (période)', 'subscriptions.new_in_period'),
        ('Revenu total (réussi)', 'revenue.total'),
        ('Revenu (période)', 'revenue.in_period'),
    ],
    'users': [
        ('Total utilisateurs', 'total'),
        ('Nouveaux utilisateurs (période)', 'new_in_period'),
    ],
    'summaries': [
        ('Total résumés générés', 'total'),
        ('Résumés générés (période)', 'in_period'),
        ('Aujourd\'hui', 'today'),
        ('Cette semaine', 'this_week'),
        ('Ce mois', 'this_month'),
    ],
    'qcm': [
        ('Total QCM générés', 'total'),
        ('  classiques', 'classic_total'),
        ('  personnalisés', 'personalized_total'),
        ('QCM générés (période)', 'in_period'),
        ('Aujourd\'hui', 'today'),
        ('Cette semaine', 'this_week'),
        ('Ce mois', 'this_month'),
    ],
    'transactions': [
        ('Total transactions (période)', 'total'),
        ('  lancées', 'by_status.pending'),
        ('  réussies', 'by_status.completed'),
        ('  échouées', 'by_status.failed'),
        ('  annulées', 'by_status.refunded'),
        ('Taux de réussite (%)', 'success_rate'),
        ('Achats de résumés (période)', 'by_type.summary.count'),
        ('  dont réussis', 'by_type.summary.completed'),
        ('  montant réussi', 'by_type.summary.amount_completed'),
        ('Abonnements (période)', 'by_type.service.count'),
        ('  dont réussis', 'by_type.service.completed'),
        ('  montant réussi', 'by_type.service.amount_completed'),
    ],
    'purchases': [
        ('Total résumés achetés (réussis)', 'total'),
        ('Achats (période)', 'in_period'),
        ('Aujourd\'hui', 'today'),
        ('Cette semaine', 'this_week'),
        ('Ce mois', 'this_month'),
        ('Montant total généré', 'amount_total'),
        ('Montant (période)', 'amount_in_period'),
        ('Montant aujourd\'hui', 'amount_today'),
        ('Montant cette semaine', 'amount_this_week'),
        ('Montant ce mois', 'amount_this_month'),
        ('Panier moyen (période)', 'avg_basket'),
        ('Panier moyen (tous temps)', 'avg_basket_all_time'),
    ],
    'subscriptions': [
        ('Abonnements actifs', 'active'),
        ('Nouveaux abonnements (période)', 'new_in_period'),
        ('Renouvelés (période)', 'renewed'),
        ('Expirés (total)', 'expired'),
        ('Expirés (période)', 'expired_in_period'),
        ('Annulés (total)', 'cancelled'),
        ('Annulés créés (période)', 'cancelled_new_in_period'),
        ('Total abonnements', 'total'),
        ('Revenus abonnements (total)', 'revenue.amount_total'),
        ('Revenus abonnements (période)', 'revenue.amount_in_period'),
        ('Revenus aujourd\'hui', 'revenue.today'),
        ('Revenus cette semaine', 'revenue.this_week'),
        ('Revenus ce mois', 'revenue.this_month'),
    ],
    'revenue': [
        ('Achats résumés (quantité total)', 'purchases.quantity'),
        ('Achats résumés (quantité période)', 'purchases.quantity_in_period'),
        ('Achats résumés (montant total)', 'purchases.amount_total'),
        ('Achats résumés (montant période)', 'purchases.amount_in_period'),
        ('Abonnements (quantité total)', 'subscriptions.quantity'),
        ('Abonnements (quantité période)', 'subscriptions.quantity_in_period'),
        ('Abonnements (montant total)', 'subscriptions.amount_total'),
        ('Abonnements (montant période)', 'subscriptions.amount_in_period'),
        ('Revenu total', 'total.amount_total'),
        ('Revenu (période)', 'total.amount_in_period'),
        ('Période précédente (montant)', 'previous_period.amount_total'),
        ('Évolution (%)', 'evolution_percent'),
    ],
}

# Séries temporelles par section : (granularité, clé valeur, libellé colonne)
_SERIES = {
    'users': ('daily', 'count', 'Nouveaux utilisateurs par jour'),
    'summaries': ('daily', 'count', 'Résumés générés par jour'),
    'qcm': ('daily', 'count', 'QCM générés par jour'),
    'transactions': ('daily', 'count', 'Transactions par jour'),
    'purchases': ('daily', 'count', 'Résumés achetés par jour'),
    'subscriptions': ('daily', 'count', 'Abonnements créés par jour'),
    'revenue': ('daily', 'total', 'Revenu par jour'),
}

_MONEY_KEYS = {
    'by_type.summary.amount_completed',
    'by_type.service.amount_completed',
    'amount_total', 'amount_in_period',
    'amount_today', 'amount_this_week', 'amount_this_month',
    'avg_basket', 'avg_basket_all_time',
    'revenue.amount_total', 'revenue.amount_in_period',
    'revenue.today', 'revenue.this_week', 'revenue.this_month',
    'purchases.amount_total', 'purchases.amount_in_period',
    'subscriptions.amount_total', 'subscriptions.amount_in_period',
    'total.amount_total', 'total.amount_in_period',
    'previous_period.amount_total',
    'revenue.total', 'revenue.in_period',
}


def _resolve(path, data):
    cur = data
    for part in path.split('.'):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur


def _collect_metrics(section, data):
    out = []
    for label, path in _METRICS.get(section, []):
        value = _resolve(path, data)
        if isinstance(value, float) and path in _MONEY_KEYS:
            value = f"{value:.2f}"
        elif value is None:
            value = 'N/A'
        out.append((label, value))
    return out


def _collect_series(section, data):
    """Séries (daily + weekly + monthly si présentes)."""
    rows = []
    if section == 'revenue':
        # Série quotidienne des revenus : achats / abonnements / total.
        for point in data.get('total', {}).get('daily', []):
            rows.append((f"daily | {point['date']} | revenu total", point.get('total')))
            rows.append((f"daily | {point['date']} | achats résumés", point.get('purchases')))
            rows.append((f"daily | {point['date']} | abonnements", point.get('subscriptions')))
        return rows
    for gran in ('daily', 'weekly', 'monthly'):
        series = data.get(gran) if isinstance(data, dict) else None
        if not series:
            continue
        # La clé de valeur réelle de la série ('count' ou 'total').
        sample = series[0] if series else {}
        val_key = 'total' if 'total' in sample else 'count'
        label = {'daily': 'par jour', 'weekly': 'par semaine', 'monthly': 'par mois'}[gran]
        for point in series:
            rows.append((f"{gran} | {point['date']} | {label}",
                         point.get(val_key)))
    return rows


def _stats_for(period, request, section):
    """Calcule les statistiques de la section (période + filtres respectés)."""
    start, end = period['start'], period['end']
    qp = request.query_params
    if section == 'users':
        return services.users_stats(start, end)
    if section == 'summaries':
        return services.summaries_stats(start, end)
    if section == 'qcm':
        return services.qcm_stats(start, end)
    if section == 'transactions':
        return services.transactions_stats(
            start, end,
            status=qp.get('status') if qp.get('status') in services.TRANSACTION_STATUSES else None,
            txn_type=qp.get('type') if qp.get('type') in ('summary', 'service') else None,
        )
    if section == 'purchases':
        return services.purchases_stats(start, end)
    if section == 'subscriptions':
        sid = _int_or_none(qp.get('service_id'))
        return services.subscriptions_stats(start, end, service_id=sid)
    if section == 'revenue':
        sid = _int_or_none(qp.get('service_id'))
        return services.revenue_stats(start, end, service_id=sid)
    return services.overview_stats(start, end)


def _int_or_none(value):
    try:
        return int(value) if value is not None else None
    except (TypeError, ValueError):
        return None


def _period_rows(period):
    return [
        ('Période', period['label']),
        ('Début', period['start'].strftime('%Y-%m-%d %H:%M')),
        ('Fin', period['end'].strftime('%Y-%m-%d %H:%M')),
    ]


def _section_content(period, request, section):
    """[(clé, valeur)] pour la section : méta période + métriques + séries."""
    data = _stats_for(period, request, section)
    rows = _period_rows(period)
    rows.append(('Section', SECTION_LABELS.get(section, section)))
    rows.append(('', ''))
    rows.extend(_collect_metrics(section, data))
    series = _collect_series(section, data)
    if series:
        rows.append(('', ''))
        rows.append(('Évolution', 'Date | Granularité | Valeur'))
        rows.extend(series)
    return rows


def build_csv_response(period, request, section):
    """Export CSV (UTF-8 avec BOM pour Excel)."""
    rows = _section_content(period, request, section)

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    for label, value in rows:
        writer.writerow([label, value])

    filename = f"statistiques_{section}_{period['start'].date()}_{period['end'].date()}.csv"
    response = HttpResponse('﻿' + buffer.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_excel_response(period, request):
    """Export Excel : une feuille par section, période et filtres respectés."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)

    for section in SECTION_LABELS:
        ws = wb.create_sheet(title=SECTION_LABELS[section][:31])
        rows = _section_content(period, request, section)
        for row_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=row_idx, column=1, value=label).font = header_font
            ws.cell(row=row_idx, column=2, value=value)
        # Largeur de colonne lisible.
        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 30

    filename = f"statistiques_{period['start'].date()}_{period['end'].date()}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
